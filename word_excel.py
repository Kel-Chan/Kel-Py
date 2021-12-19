from multiprocessing import Value
import os
from typing import ItemsView
import docx
from numpy.core.numerictypes import maximum_sctype
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from glob import glob
from multiprocessing import Process
import xlsxwriter


i=0
for file in os.listdir('.'):
    if os.path.isfile(file) and file.split('.')[1] == 'docx': #file以'.'分開，如果'.'後面是'pdf'就進下一步
        wordName=file
        doc = docx.Document(wordName)
        i=i+1
        print("正在處理文件{0}".format(wordName))
        
        baseName = wordName.split('.')[0]
        excelName = baseName + '.xlsx'
        a = doc.paragraphs[0].text.replace(",","")
        b = np.array(a.split())

        eq_letter_1 = np.where(b=="1")[0][0]
        eq_letter_10 = np.where(b=="10")[0][-1]
        eq_letter_20 = np.where(b=="20")[0][-1]
        eq_letter_30 = np.where(b=="30")[0][-1]

        Premium_locate = eq_letter_1 + 1
        Guar_10_locate = eq_letter_10 + 2
        Guar_20_locate = eq_letter_20 + 2
        Guar_30_locate = eq_letter_30 + 2

        NonGuar_10_locate_add4 = eq_letter_10 + 4
        NonGuar_20_locate_add4 = eq_letter_20 + 4
        NonGuar_30_locate_add4 = eq_letter_30 + 4

        NonGuar_10_locate_add5 = eq_letter_10 + 5
        NonGuar_20_locate_add5 = eq_letter_20 + 5
        NonGuar_30_locate_add5 = eq_letter_30 + 5

        Premium = int(b[Premium_locate])
        Guar_10 = int(b[Guar_10_locate])
        Guar_20 = int(b[Guar_20_locate])
        Guar_30 = int(b[Guar_30_locate])

        if int(b[NonGuar_10_locate_add4])>int(b[NonGuar_10_locate_add5]):
            NonGuar_10 = int(b[NonGuar_10_locate_add4])
        else:
            NonGuar_10 = int(b[NonGuar_10_locate_add5])
        if int(b[NonGuar_20_locate_add4])>int(b[NonGuar_20_locate_add5]):
            NonGuar_20 = int(b[NonGuar_20_locate_add4])
        else:
            NonGuar_20 = int(b[NonGuar_20_locate_add5])
        if int(b[NonGuar_30_locate_add4])>int(b[NonGuar_30_locate_add5]):
            NonGuar_30 = int(b[NonGuar_30_locate_add4])
        else:
            NonGuar_30 = int(b[NonGuar_30_locate_add5])
        

        items = [[baseName],[Premium],[Guar_10],[NonGuar_10],[Guar_20],[NonGuar_20],[Guar_30],[NonGuar_30]]
        with xlsxwriter.Workbook(excelName) as workbook:
            worksheet = workbook.add_worksheet('report')
            for row_num, data in enumerate(items,0):
                    worksheet.write_row(row_num,1,data)
                
        print(i)
        print("<{0}>第{1}份處理完成!".format(baseName,i))
        print("{0}處理完成,保存為{1}!!!!".format(wordName,excelName))


for file in os.listdir('.'):
    if os.path.isfile(file) and file.split('.')[1] == 'xlsx': #file以'.'分開，如果'.'後面是'pdf'就進下一步
        excelName=file
        print("正在處理文件{0}".format(excelName))
        baseName = excelName.split('.')[0]
        i=i+1 
        items = ['檔案名稱'],["每年保費"],["保證回報(10)"],["預期回報(10)"],["保證回報(20)"],["預期回報(20)"],["保證回報(30)"],["保證回報(30)"]
        wb=load_workbook(excelName)
        ws=wb.active
        ws['A1'] = '檔案名稱'
        ws['A2'] = '每年保費'
        ws['A3'] = '保證回報(10)'
        ws['A4'] = '預期回報(10)'
        ws['A5'] = '保證回報(20)'
        ws['A6'] = '預期回報(20)'
        ws['A7'] = '保證回報(30)'
        ws['A8'] = '預期回報(30)'

        wb.save(excelName)
        print("<{0}>第{1}份處理完成!".format(excelName,i))
