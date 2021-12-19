import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

i=0
for file in os.listdir('.'):
    if os.path.isfile(file) and file.split('.')[1] == 'xlsx': #file以'.'分開，如果'.'後面是'pdf'就進下一步
        excelName=file
        print("正在處理文件{0}".format(excelName))
        baseName = excelName.split('.')[0]
        i=i+1 
        items = ['檔案名稱'],["每年保費"],["保證回報(10)"],["預期回報(10)"],["保證回報(20)"],["預期回報(20)"],["保證回報(30)"],["保證回報(30)"]
        wb=load_workbook(excelName)
        ws=wb.active
        ws['A1'].str = '檔案名稱'
        ws['A2'].str = '每年保費'
        ws['A3'].str = '保證回報(10)'
        ws['A4'].str = '預期回報(10)'
        ws['A5'].str = '保證回報(20)'
        ws['A6'].str = '預期回報(20)'
        ws['A7'].str = '保證回報(30)'
        ws['A8'].str = '預期回報(30)'

        wb.save(excelName)
        print("<{0}>第{1}份處理完成!".format(excelName,i))
